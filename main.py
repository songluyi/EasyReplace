# -*- coding: utf-8 -*-
# 2019/4/19 09:31
"""
-------------------------------------------------------------------------------
Function:   EasyReplace Main Function
Version:    1.0
Author:     SLY
Contact:    slysly759@gmail.com
code is far away from bugs with the god Animal protecting
               ┏┓      ┏┓
            ┏┛┻━━━┛┻┓
            ┃      ☃      ┃
            ┃  ┳┛  ┗┳  ┃
            ┃      ┻      ┃
            ┗━┓      ┏━┛
                ┃      ┗━━━┓
                ┃  神兽保佑    ┣┓
                ┃　永无BUG！   ┏┛
                ┗┓┓┏━┳┓┏┛
                  ┃┫┫  ┃┫┫
                  ┗┻┛  ┗┻┛

-------------------------------------------------------------------------------
"""

import sys, os
import pickle
import re
import codecs
import string
import shutil
from win32com import client as wc
# import docx
from openpyxl import load_workbook
import datetime
from time import sleep
import pythoncom

def ChangeDoc2Docx(FilePath):
    word = wc.Dispatch('Word.Application')
    doc = word.Documents.Open(FilePath)  # 目标路径下的文件
    NewPath = FilePath + 'x'
    doc.SaveAs(NewPath, 12, False, "", True, "", False, False, False, False)  # 转化后路径下的文件
    doc.Close()
    word.Quit()

# def highlight_all(word_file,find_str):
#     ''' replace all occurrences of `find_str` w/ `replace_str` in `word_file` '''
#     wdFindContinue = 1
#     wdReplaceAll = 2
#
#     # Dispatch() attempts to do a GetObject() before creating a new one.
#     # DispatchEx() just creates a new one.
#     app = wc.DispatchEx("Word.Application")
#     app.Visible = True
#     app.DisplayAlerts = False
#     app.Documents.Open(word_file)
#     app = wc.DispatchEx("Word.Application")
#     app.Visible = False
#     app.DisplayAlerts = False
#     app.Documents.Open(word_file)
#     app.Selection.Find.HitHighlight(find_str,HighlightColor=32896,MatchCase=False,MatchWholeWord=False,MatchWildcards=False,
#                                     MatchSoundsLike=False,MatchAllWordForms=False,IgnoreSpace=True)
#     app.ActiveDocument.Close(SaveChanges=True)
#     app.Quit()
#     return 1

def search_replace_all(word_file, ReplaceDict):
    ''' replace all occurrences of `find_str` w/ `replace_str` in `word_file` '''
    wdFindContinue = 1
    wdReplaceAll = 2

    # Dispatch() attempts to do a GetObject() before creating a new one.
    # DispatchEx() just creates a new one.
    app = wc.DispatchEx("Word.Application")
    app.Visible = True
    app.DisplayAlerts = False
    doc=app.Documents.Open(word_file)
    for form in ReplaceDict:


        # expression.HitHighlight(FindText, HighlightColor, TextColor, MatchCase, MatchWholeWord, MatchPrefix,
        #                         MatchSuffix, MatchPhrase, MatchWildcards, MatchSoundsLike, MatchAllWordForms, MatchByte,
        #                         MatchFuzzy, MatchKashida, MatchDiacritics, MatchAlefHamza, MatchControl, IgnoreSpace,
        #                         IgnorePunct, HanjaPhoneticHangul)
        # app.Selection.Find.HitHighlight(str(form[0]),HighlightColor=32896,MatchCase=False,MatchWholeWord=True,
        #                                 MatchWildcards=False,MatchSoundsLike=False,MatchAllWordForms=False)
        # 高亮之后没有保存的任何效果。。。
        # with app.ActiveDocument.Content.Find(form[0]):

        # app.Selection.Find.HitHighlight(str(form[0]),HighlightColor=16711680)
        print(str(form[0]) + '||替换为==》||' + str(form[1]))
        flag=app.Selection.Find.Execute(str(form[0]), False, False, False, False, False,
                                   True, wdFindContinue, False, str(form[1]), wdReplaceAll)

    # for form in ReplaceDict:
    #     print(str(form[0])+'||高亮为==》||'+str(form[1]))
    #     app.Selection.Find.HitHighlight(str(form[1]),HighlightColor=16711680)
    try:
        app.ActiveDocument.Close(SaveChanges=True)
        app.Application.Quit(-1)
        # del (app)
    except Exception as e:
        print(e)

    return 1


def NewReplaceWord(WordList, ReplaceDict):
    # for f in WordList:
    #     # doc=word.Documents.Open(f)
    #     for From in ReplaceDict.values():
    #         highlight_all(f, str(From[0]))
    for f in WordList:
        search_replace_all(f, ReplaceDict.values())



# 该函数无法解决替换后格式问题
# def ReplaceWord(WordList, ReplaceDict):
#     i = 0
#     if (len(WordList) > 0):
#         print("正在读取word文档进行替换中")
#
#         for f in WordList:
#             document = docx.Document(f)
#             for p in document.paragraphs:
#                 for value in ReplaceDict.values():
#                     ## 先做一个检测 是否存在 需要替换的关键词 在做替换加快效率
#                     TEST = p.text
#                     if str(value[0]) in p.text:
#                         print(str(value[0]) + 'found!')
#                         # 这种写法没办法保留格式
#
#                         inline = p.runs
#                         if i in range(len(inline)):
#                             ## 这种样式必须保持一致才可以 不过我有一个大胆的想法
#                             if str(value[0]) in inline[i].text:
#                                 print(str(value[0]) + '||替换--》' + inline[i].text)
#                                 text = inline[i].text.replace(str(value[0]), str(value[1]))
#                                 inline[i].text = text
#                         p.text = p.text.replace(str(value[0]), str(value[1]))
#             document.save(f)
#
#     return


def GetExcelDate():
    ReplaceDict = {}
    # 如果替换词与被替换词检测一致不放在dict 里面
    name = 'setting.xlsx'
    excel_path = os.getcwd() + '\\' + 'data\\' + name
    wb_get_excel = load_workbook(filename=excel_path, data_only=True)
    sheets = wb_get_excel.sheetnames
    ws_get_excel = wb_get_excel[sheets[0]]
    line = 2
    while ws_get_excel.cell(row=line, column=1).value is not None:
        DictKey = ws_get_excel.cell(row=line, column=2).value
        # 针对日期格式数据还要做一轮新转换
        if type(ws_get_excel.cell(row=line, column=3).value) is datetime.datetime:
            NowRowTime = ws_get_excel.cell(row=line, column=3).value
            NowTime = str(NowRowTime.year) + '年' + str(NowRowTime.month) + '月' + str(NowRowTime.day) + "日"
            ReplaceRowTime = ws_get_excel.cell(row=line, column=4).value
            ReplaceTime = str(ReplaceRowTime.year) + '年' + str(ReplaceRowTime.month) + '月' + str(
                ReplaceRowTime.day) + "日"
            DictValue = [NowTime, ReplaceTime]
        else:
            DictValue = [ws_get_excel.cell(row=line, column=3).value, ws_get_excel.cell(row=line, column=4).value]
        if (ws_get_excel.cell(row=line, column=3).value != ws_get_excel.cell(row=line, column=4).value):
            ReplaceDict[DictKey] = DictValue
        line += 1
    print(ReplaceDict)
    return ReplaceDict


def get_path():
    current_path = os.getcwd()
    new_path = current_path + '\\' + 'data\\'
    FileList = []
    rootdir = new_path

    for root, subFolders, files in os.walk(rootdir):
        for f in files:
            if f.find('doc') != -1:
                FileList.append(os.path.join(root, f))

    print('检测到您目录下有如下world文档 请确保他们是要批量替换的文件')
    for item in FileList:
        print(item)
    return FileList


if __name__ == '__main__':
    WordList = get_path()
    Num=input('请按任意键回车确认继续')
    if Num is not None:
        ExcelDict = GetExcelDate()
        NewReplaceWord(WordList, ExcelDict)
        print("替换完成请注意查阅目录文档")